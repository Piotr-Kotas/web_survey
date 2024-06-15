# app.py

import sqlite3
import os
import openpyxl
from flask import Flask, render_template, request, redirect, url_for, session, g, jsonify
from openpyxl.utils import get_column_letter
from populate_queries import populate_queries
from pyutil import reload_queries, reload_user_answers, reload_manager_answers, query_size, query_string
DATABASE = './data/site.db'


def generate_excel_export():
    # Connect to the SQLite database
    conn = sqlite3.connect('data/site.db')
    cursor = conn.cursor()

    # Retrieve user answers
    cursor.execute('SELECT * FROM useranswers')
    user_answers = cursor.fetchall()

    # Retrieve manager answers
    cursor.execute('SELECT * FROM manageranswers')
    manager_answers = cursor.fetchall()

    # Retrieve questions
    cursor.execute('SELECT question FROM queries')
    questions = cursor.fetchall()

    # Create Excel workbook
    workbook = openpyxl.Workbook()

    # Add user answers to first sheet
    user_sheet = workbook.active
    user_sheet.title = 'User Answers'
    headers = ['Username'] + [f'Q{i}' for i in range(1, len(questions) + 1)]
    user_sheet.append(headers)
    for answer in user_answers:
        user_sheet.append([answer[1]] + list(answer[2:]))

    # Add manager answers to second sheet
    manager_sheet = workbook.create_sheet(title='Manager Answers')
    headers = ['Managername', 'Username'] + [f'Q{i}' for i in range(1, len(questions) + 1)]
    manager_sheet.append(headers)
    for answer in manager_answers:
        manager_sheet.append([answer[1], answer[2]] + list(answer[3:]))

    # Add evaluation to third sheet
    eval_sheet = workbook.create_sheet(title='Evaluation')
    eval_sheet.append(['Username'] + [f'Q{i}' for i in range(1, len(questions) + 1)])
    for user_answer, manager_answer in zip(user_answers, manager_answers):
        eval_row = [user_answer[1]]
        for i in range(1, len(questions) + 1):
            user_response = user_answer[i + 1]
            manager_response = manager_answer[i + 2]
            if user_response == 'I can':
                eval_row.append('Y')
            elif user_response == "I can't" and manager_response == "I can't":
                eval_row.append('N')
            elif user_response == "I'll learn" and (manager_response == "I'll learn" or manager_response == "I can't"):
                eval_row.append('L')
            elif user_response == "I can't" and (manager_response == "I'll learn" or manager_response == "I can"):
                eval_row.append('L')
            elif user_response == "I'll learn" and manager_response == "I can":
                eval_row.append('Y')
        eval_sheet.append(eval_row)

    # Save workbook
    workbook.save('data/evaluation.xlsx')

    # Close database connection
    conn.close()

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DATABASE)
    return db

def query_db(query, args=(), one=False):
    cur = get_db().execute(query, args)
    rv = cur.fetchall()
    cur.close()
    return (rv[0] if rv else None) if one else rv

app = Flask(__name__)
app.secret_key = 'supersecretkey'
Qstring=""
SignString=""
CreateString=""

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

@app.route('/export_excel', methods=['POST'])
def export_excel():
    # Call the function to export data to Excel
    # Replace generate_excel_export() with your actual function
    generate_excel_export()
    # Return a JSON response
    return jsonify({'message': 'Data exported to Excel successfully'})

def init_db():
    with app.app_context():
        db = get_db()
        with app.open_resource('schema.sql', mode='r') as f:
            db.cursor().executescript(f.read())
        db.commit()

@app.route('/initdb')
def initialize_db():
    init_db()
    return 'Database Initialized'

@app.route('/populate_queries')
def initialize_queries():
    populate_queries()
    return 'Queries Table Populated'

@app.route('/')
def landing():
    return render_template('landing.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = query_db('SELECT * FROM credentials WHERE username = ? AND password = ?', [username, password], one=True)
        if user:
            session['username'] = user[1]
            session['role'] = user[3]
            if user[3] == 'admin':
                return redirect(url_for('admin_logged_in'))
            elif user[3] == 'manager':
                return redirect(url_for('manager_logged_in'))
            else:
                return redirect(url_for('user_logged_in'))
        else:
            return 'Invalid credentials'
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        role = request.form['role']
        db = get_db()
        try:
            db.execute('INSERT INTO credentials (username, password, role) VALUES (?, ?, ?)', (username, password, role))
            db.commit()
            return redirect(url_for('login'))
        except sqlite3.IntegrityError:
            return 'Username already exists'
    return render_template('register.html')

@app.route('/user_logged_in', methods=['GET', 'POST'])
def user_logged_in():
    if 'username' in session and session['role'] == 'user':
        questions = query_db('SELECT question FROM queries')
        if request.method == 'POST':
            answers = [request.form.get(f'q{i}') for i in range(1, len(questions) + 1)]
            db = get_db()
            db.execute(f'REPLACE INTO useranswers (username{Qstring}) VALUES (?,{SignString})',[session['username']] + answers)
            db.commit()
            return render_template('user_logged_in.html', questions=questions, success=True)
        return render_template('user_logged_in.html', questions=questions, success=False)
    return redirect(url_for('login'))


@app.route('/admin_logged_in')
def admin_logged_in():
    if 'username' in session and session['role'] == 'admin':
        return render_template('admin_logged_in.html')
    return redirect(url_for('login'))

@app.route('/manager_logged_in', methods=['GET', 'POST'])
def manager_logged_in():
    if 'username' in session and session['role'] == 'manager':
        questions = query_db('SELECT question FROM queries')
        users = query_db('SELECT username FROM credentials WHERE role="user"')
        if request.method == 'POST':
            username = request.form['username']
            answers = [request.form.get(f'q{i}') for i in range(1, len(questions) + 1)]
            db = get_db()
            # Check if the manageranswers table has a row with the same username
            existing_row = query_db('SELECT * FROM manageranswers WHERE username = ?', [username], one=True)
            if existing_row:
                # If the row exists, delete it
                db.execute('DELETE FROM manageranswers WHERE id = ?', [existing_row[0]])
            db.execute(f'REPLACE INTO manageranswers (managername, username{Qstring}) VALUES (?,?,{SignString})', [session['username'], username] + answers)
            db.commit()
            
            return render_template('manager_logged_in.html', questions=questions, users=[user[0] for user in users], success=True)
        return render_template('manager_logged_in.html', questions=questions, users=[user[0] for user in users], success=False)
    return redirect(url_for('login'))
@app.route('/admin_register', methods=['GET', 'POST'])
def admin_register():
    if request.method == 'POST':
        # Handle form submission
        # Extract form data and perform registration logic
        username = request.form['username']
        password = request.form['password']
        role = request.form['role']
        # Perform registration logic
        db = get_db()
        try:
            db.execute('INSERT INTO credentials (username, password, role) VALUES (?, ?, ?)', (username, password, role))
            db.commit()
        except sqlite3.IntegrityError:
            return 'Username already exists'
        # Assuming successful registration, show a success message
        notification = f"User {username} added successfully."
        return render_template('admin_register.html', notification=notification)
    else:
        # Display the registration form
        return render_template('admin_register.html')
    
@app.route('/user_management')
def user_management():
    # Check if the user is logged in and has the role of 'admin'
    if 'username' in session and session['role'] == 'admin':
        # Fetch the users from the database
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        cursor.execute("SELECT username, password, role FROM credentials")
        users = cursor.fetchall()
        conn.close()
        #print(users)
        # Render the template with the users data
        return render_template('user_management.html', users=users)
    else:
        # Redirect unauthorized users to the login page
        return redirect(url_for('login'))
    
@app.route('/update_user', methods=['GET', 'POST'])
def update_user():
    username = request.form['username']
    new_password = request.form['newpassword']
    role = request.form['role']

    # Update the user in the database
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    print(f"trying to update {username}")
    if not new_password:
        cursor.execute("SELECT password FROM credentials WHERE username = ?", (username,))
        old_password = cursor.fetchone()[0]   
        new_password = old_password  # Use the old password
        cursor.fetchall()
    cursor.execute("UPDATE credentials SET password = ?, role = ? WHERE username = ?", (new_password, role, username))
    conn.commit()
    conn.close()

    # Redirect back to the user management page
    return redirect(url_for('user_management'))

@app.route('/delete_user', methods=['GET','POST'])
def delete_user():
    username = request.json.get('username')  # Fetch username from request JSON
    print(f"trying to delete {username}")
    # Add your database deletion logic here
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM credentials WHERE username = ?", (username,))
    conn.commit()
    conn.close()
    return redirect(url_for('user_management'))  # Redirect back to the user management page

@app.route('/reload_tables', methods=['POST'])
def reload_tables():
    if 'username' in session and session['role'] == 'admin':
        if request.method == 'POST':
            reload_queries()
            Qstring, SignString, CreateString = query_string()
            reload_user_answers(CreateString)
            reload_manager_answers(CreateString)
        return redirect(url_for('admin_logged_in'))
    return redirect(url_for('login'))

@app.route('/logout', methods=['GET', 'POST'])
def logout():
    session.clear()
    return redirect(url_for('landing'))

if __name__ == '__main__':
    Qstring,SignString,CreateString=query_string()
    if not os.path.exists('./data/site.db'):
        os.makedirs('./data', exist_ok=True)
        init_db()
        populate_queries()
        reload_tables()    
    app.run(debug=True)
